import csv
import glob
import tkinter as tk
import dateutil.parser
import re
import os
from tkinter import filedialog

from openpyxl import Workbook
from spire.xls.common import *

from classification import classification

date_time1 = []  # 交易时间
transaction_type1 = []  # 交易类型
store1 = []  # 商家名称
commodity1 = []  # 商品名称
in_or_out1 = []  # 收支
money1 = []  # 金额
Payment_method1 = []  # 支付方式
category1 = []  # 一级分类
subclass1 = []  # 二级分类

# ali_or_wechat =[]

yimu = Workbook()
ym = yimu.active

root = tk.Tk()
root.withdraw()
folder_path = filedialog.askdirectory()
path1 = folder_path + "/DelLoad.csv"
try:
    os.remove(path1)
except OSError:
    pass
file_path = folder_path + "/*.csv"
csv1 = glob.glob(folder_path + "/*.csv")  # 获取文件夹内csv文件


for i in range(0, len(csv1)):
    len1 = []
    date_time = []
    in_or_out = []
    store = []
    category = []
    subclass = []
    money = []
    Payment_method = []
    transaction_type = []
    commodity = []
    try:
        with open(csv1[i], 'r', encoding='utf-8') as file1:
            reader = csv.reader(file1)
            for row in reader:
                len1.append(row)
    except UnicodeDecodeError:
        with open(csv1[i], newline='', encoding='ansi') as file1:
            reader = csv.reader(file1)
            for row in reader:
                len1.append(row)
    # file = os.path.splitext(os.path.basename(csv1[i]))[0]  # 拆出文件名

    if "alipay" in csv1[i]:
        del (len1[0:24])
    elif "微信" in csv1[i]:
        del (len1[0:16])
    len2 = []
    # 支付宝['交易时间', '交易分类', '交易对方', '对方账号', '商品说明', '收/支', '金额', '收/付款方式', '交易状态', '交易订单号', '商家订单号', '备注']
    # 微信['交易时间', '交易类型', '交易对方', '商品', '收/支', '金额(元)', '支付方式', '当前状态', '交易单号', '商户单号', '备注']
    for b in range(1, len(len1)):

        len2 = len1[b]

        date_time.append(dateutil.parser.parse(len2[0]))
        transaction_type.append(len2[1])
        store.append(len2[2])

        if "alipay" in csv1[i]:
            in_or_out.append(len2[5])
            commodity.append(len2[4])
            money.append(len2[6])
            Payment_method.append(len2[7])

        elif "微信" in csv1[i]:
            in_or_out.append(len2[4])
            commodity.append(len2[3])
            money.append(len2[5])
            Payment_method.append(len2[6])
    print(date_time)
    len_date_time = len(date_time)
    print(len_date_time)
    y = 0
    print(in_or_out)
    while y < len_date_time:
        print(y)
        if in_or_out[y] == "不计收支":
            date_time.pop(y)
            in_or_out.pop(y)
            commodity.pop(y)
            money.pop(y)
            Payment_method.pop(y)
            transaction_type.pop(y)
            store.pop(y)
            len_date_time = len(date_time)

        else:
            if "alipay" in csv1[i]:
                if Payment_method[y] == "/":
                    Payment_method[y] = "支付宝"
                cc = classification(in_or_out[y], commodity[y], store[y])  # 收支，商品，交易对方
                if cc is not None:
                    str1, str2 = cc
                    category.append(str1)
                    subclass.append(str2)
            elif "微信" in csv1[i]:
                if Payment_method[y] == "零钱" or Payment_method[y] == "/":
                    Payment_method[y] = "微信钱包"
                cc = classification(in_or_out[y], store[y], transaction_type[y])  # 收支，商店名字，交易类型，
                if cc is not None:
                    str1, str2 = cc
                    category.append(str1)
                    subclass.append(str2)
            y += 1
date_time1 += date_time
in_or_out1 += in_or_out
store1 += store
category1 += category
subclass1 += subclass

money1 += money
Payment_method1 += Payment_method
commodity1 += commodity
transaction_type1 += transaction_type


for y in range(0, len(date_time1)):  # 如果不需要分大餐，请将这部分代码注释
    for s in re.findall(r"-?\d+\.?\d*", money1[y]):
        float_s = float(s)
        ym.cell(y + 2, 3).value = float_s  # 金额
    if subclass1[y] == "三餐" and float_s > 20:
        subclass1[y] = "大餐"

    ym.cell(y + 2, 1).value = date_time1[y]  # 时间
    ym.cell(y + 2, 2).value = in_or_out1[y]  # 收支
    ym.cell(y + 2, 4).value = category1[y]  # 一级分类
    ym.cell(y + 2, 5).value = subclass1[y]  # 二级分类
    ym.cell(y + 2, 6).value = "日常账本"
    ym.cell(y + 2, 8).value = "{0}-{1}".format(commodity1[y], store1[y])
    ym.cell(y + 2, 7).value = Payment_method1[y]

list1 = ['日期', '收支类型', '金额', '类别', '子类', '所属账本', '收支账户', '备注']
for i in range(0, len(list1)):
    ym.cell(1, i+1).value = list1[i]

try:
    os.remove(path1)
    os.makedirs(folder_path + "/完成")
except OSError:
    pass
yimu_name = folder_path + "/完成/sc.xlsx"
yimu.save(yimu_name)
os.startfile(folder_path + "/完成")
