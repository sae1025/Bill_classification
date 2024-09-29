import csv
import glob
import tkinter as tk
import shutil
import re
from tkinter import filedialog

from openpyxl import Workbook
from spire.xls.common import *

from 分类函数 import classification

date_time1 = []  # 交易时间
transaction_type1 = []  # 交易类型
store1 = []  # 商家名称
commodity1 = []  # 商品名称
in_or_out1 = []  # 收支
money1 = []  # 金额
Payment_method1 = []  # 支付方式
category1 = []  # 一级分类
subclass1 = []  # 二级分类

# ali_or_wechat =[]

yimu = Workbook()
ym = yimu.active

root = tk.Tk()
root.withdraw()
folder_path = filedialog.askdirectory()
path1 = folder_path + "/DelLoad.csv"
try:
    os.remove(path1)
except OSError:
    pass
file_path = folder_path + "/*.csv"
csv1 = glob.glob(folder_path + "/*.csv")  # 获取文件夹内csv文件


for i in range(0, len(csv1)):
    len1 = []
    date_time = []
    in_or_out = []
    store = []
    category = []
    subclass = []
    money = []
    Payment_method = []
    transaction_type = []
    commodity = []
    try:
        with open(csv1[i], 'r', encoding='utf-8') as file1:
            reader = csv.reader(file1)
            for row in reader:
                len1.append(row)
    except UnicodeDecodeError:
        with open(csv1[i], newline='', encoding='ansi') as file1:
            reader = csv.reader(file1)
            for row in reader:
                len1.append(row)
    # file = os.path.splitext(os.path.basename(csv1[i]))[0]  # 拆出文件名

    if "alipay" in csv1[i]:
        del (len1[0:24])
    if "微信" in csv1[i]:
        del (len1[0:16])

    with open(path1, 'w', newline='', encoding='utf-8') as file:
        csv_writer = csv.writer(file)
        csv_writer.writerows(len1)
    with open(path1, 'r', encoding='utf-8') as file1:
        reader = csv.DictReader(file1)
        for row in reader:
            date_time.append(row['交易时间'])
            in_or_out.append(row["收/支"])
            store.append(row["交易对方"])
            if "alipay" in csv1[i]:
                commodity.append(row['商品说明'])
                money.append(row["金额"])
                Payment_method.append(row['收/付款方式'])
                transaction_type.append(row['交易分类'])
            if "微信" in csv1[i]:
                commodity.append(row['商品'])
                transaction_type.append(row['交易类型'])
                money.append(row["金额(元)"])
                Payment_method.append(row['支付方式'])
    len_date_time = len(date_time)
    y = 0
    while y < len(date_time):
        if in_or_out[y] == "不计收支":
            date_time.pop(y)
            in_or_out.pop(y)
            commodity.pop(y)
            money.pop(y)
            Payment_method.pop(y)
            transaction_type.pop(y)
            store.pop(y)
            len_date_time = len(date_time)
        else:
            if "alipay" in csv1[i]:
                cc = classification(in_or_out[y], commodity[y], store[y])  # 收支，商品，交易对方
                if Payment_method[y] == "/":
                    Payment_method[y] = "支付宝"
            if "微信" in csv1[i]:
                if Payment_method[y] == "零钱" or Payment_method[y] == "/":
                    Payment_method[y] = "微信钱包"
                cc = classification(in_or_out[y], store[y], transaction_type[y])  # 收支，商店名字，交易类型，

            if cc is not None:
                str1, str2 = cc
                category.append(str1)
                subclass.append(str2)
            else:
                pass
            y += 1
    date_time1 += date_time
    in_or_out1 += in_or_out
    store1 += store
    category1 += category
    subclass1 += subclass
    money1 += money
    Payment_method1 += Payment_method
    commodity1 += commodity
    transaction_type1 += transaction_type


for y in range(0, len(date_time)):
    for s in re.findall(r"-?\d+\.?\d*", money[y]):
        float_s = float(s)
        ym.cell(y + 2, 3).value = float_s  # 金额
    if subclass[y] == "三餐" and float_s > 20:
        ym.cell(y + 2, 5).value = "大餐"  # 二级分类

    ym.cell(y + 2, 1).value = date_time1[y]  # 时间
    ym.cell(y + 2, 2).value = in_or_out1[y]  # 收支
    ym.cell(y + 2, 4).value = category1[y]  # 一级分类
    ym.cell(y + 2, 5).value = subclass1[y]  # 二级分类
    ym.cell(y + 2, 6).value = "日常账本"
    ym.cell(y + 2, 8).value = "{0}-{1}".format(commodity1[y], store1[y])
    ym.cell(y + 2, 7).value = Payment_method1[y]

list1 =['日期', '收支类型', '金额', '类别', '子类', '所属账本', '收支账户', '备注']
for i in (0, len(list1)):
    ym.cell(1, i+1).value = list1[i]
try:
    os.remove(path1)
    os.makedirs(folder_path + "/完成")
except OSError:
    pass
yimu_name = folder_path + "/完成/sc.xlsx"
yimu.save(yimu_name)


